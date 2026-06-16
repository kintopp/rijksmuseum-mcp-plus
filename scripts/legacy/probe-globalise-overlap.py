"""Pre-Stage-E sizing probe for GLOBALISE/ESTA gazetteer overlap with the
v0.25 cold-rerun vocabulary DB. Read-only.

Produces:
- single-line summary on stdout
- per-row CSV at <output_dir>/candidates.csv with columns
  vocab_id,label,db_lat,db_lon,source,src_id,src_name,src_lat,src_lon,
  src_class,distance_km,bucket

Decision rule (per offline/v0.25-integration-report.md §4):
- ungeocoded_match >= 30  → schedule phase 1f for v0.26
- ungeocoded_match  < 30  → label-enrichment only; skip new phase

Run:
  ~/miniconda3/envs/embeddings/bin/python scripts/probe-globalise-overlap.py \\
    --db data/vocabulary.db \\
    --esta '/Users/abosse/Documents/GitHub/globalise/dataverse/ESTA locations/ESTA_standardised_locations_2025_12_18_v7.xlsx' \\
    --globalise '/Users/abosse/Documents/GitHub/globalise/dataverse/Globalise places/GLOBALISE - Places in the Dutch East India Company Archives (1602-1799).xlsx' \\
    --output-dir data/audit/globalise-probe/2026-04-29/
"""
from __future__ import annotations

import argparse
import csv
import math
import sqlite3
from collections import defaultdict
from pathlib import Path

import openpyxl


def normalise(name: str | None) -> str:
    if not name:
        return ""
    return str(name).strip().lower()


def to_float(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    r = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


def load_esta(path: Path) -> dict[str, list[dict]]:
    """name_map: normalised name -> list of match dicts."""
    name_map: dict[str, list[dict]] = defaultdict(list)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Pass 1: STANDARDISED_SUMMARY → canonical names with coords
    summary_by_id: dict[str, dict] = {}
    ws = wb["STANDARDISED_SUMMARY"]
    rows = ws.iter_rows(min_row=2, values_only=True)
    for r in rows:
        loc_id = r[0]
        name = r[1]
        lat, lon = to_float(r[16]), to_float(r[17])
        exact_class = (r[18] or "").strip().lower()
        if not loc_id or not name:
            continue
        rec = {"source": "esta", "src_id": str(loc_id), "src_name": str(name),
               "src_lat": lat, "src_lon": lon, "src_class": exact_class}
        summary_by_id[str(loc_id)] = rec
        name_map[normalise(name)].append(rec)

    # Pass 2: STANDARDISED_NAME_ORIGINAL → spelling variants pointing at same coords
    ws = wb["STANDARDISED_NAME_ORIGINAL"]
    rows = ws.iter_rows(min_row=2, values_only=True)
    for r in rows:
        loc_id = r[0]
        variant = r[1]
        if not loc_id or not variant:
            continue
        parent = summary_by_id.get(str(loc_id))
        if not parent:
            continue
        n = normalise(variant)
        if parent not in name_map[n]:
            name_map[n].append(parent)

    return name_map


def load_globalise(path: Path) -> dict[str, list[dict]]:
    name_map: dict[str, list[dict]] = defaultdict(list)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Sheet2: pref + alt labels with canonical coords
    overview_by_id: dict[str, dict] = {}
    ws = wb["Sheet2 Places – Overview"]
    rows = ws.iter_rows(min_row=2, values_only=True)
    for r in rows:
        glob_id = r[0]
        pref = r[1]
        alts = (r[2] or "").split("|") if r[2] else []
        lat, lon = to_float(r[4]), to_float(r[5])
        if not glob_id or not pref:
            continue
        rec = {"source": "globalise", "src_id": str(glob_id), "src_name": str(pref),
               "src_lat": lat, "src_lon": lon, "src_class": ""}
        overview_by_id[str(glob_id)] = rec
        name_map[normalise(pref)].append(rec)
        for alt in alts:
            n = normalise(alt)
            if n and rec not in name_map[n]:
                name_map[n].append(rec)

    # Sheet4: full label list (PREF/ALT) — broader than Sheet2 alt_labels
    ws = wb["Sheet4 Places – Labels Detail"]
    rows = ws.iter_rows(min_row=2, values_only=True)
    for r in rows:
        glob_id = r[0]
        label = r[2]
        if not glob_id or not label:
            continue
        parent = overview_by_id.get(str(glob_id))
        if not parent:
            continue
        n = normalise(label)
        if n and parent not in name_map[n]:
            name_map[n].append(parent)

    return name_map


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", required=True)
    ap.add_argument("--esta", required=True)
    ap.add_argument("--globalise", required=True)
    ap.add_argument("--output-dir", required=True)
    ap.add_argument("--disagreement-km", type=float, default=50.0)
    args = ap.parse_args()

    out = Path(args.output_dir)
    out.mkdir(parents=True, exist_ok=True)

    esta = load_esta(Path(args.esta))
    glob = load_globalise(Path(args.globalise))
    print(f"Loaded ESTA: {len(esta):,} normalised name keys")
    print(f"Loaded GLOBALISE: {len(glob):,} normalised name keys")

    conn = sqlite3.connect(args.db)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT id, label_en, label_nl, lat, lon FROM vocabulary "
        "WHERE type='place'"
    ).fetchall()
    print(f"Loaded {len(rows):,} place rows from DB")

    # Bucketing
    counts = {
        "ungeocoded_match": 0,             # DB lat NULL, gazetteer match found
        "ungeocoded_unique_vocab": 0,      # distinct vocab IDs in above
        "ungeocoded_no_match": 0,
        "geocoded_match": 0,
        "geocoded_disagree": 0,
        "geocoded_match_no_src_coord": 0,
    }
    matched_vocab_ids: set[str] = set()
    csv_rows: list[list] = []

    for row in rows:
        vid, label_en, label_nl, db_lat, db_lon = row
        labels = [normalise(label_en), normalise(label_nl)]
        labels = [l for l in labels if l]
        primary_label = label_en or label_nl or ""

        hits: list[dict] = []
        seen_keys: set[tuple[str, str]] = set()
        for n in labels:
            for m in esta.get(n, []) + glob.get(n, []):
                key = (m["source"], m["src_id"])
                if key in seen_keys:
                    continue
                seen_keys.add(key)
                hits.append(m)

        if db_lat is None:
            if hits:
                counts["ungeocoded_match"] += len(hits)
                matched_vocab_ids.add(vid)
                for m in hits:
                    csv_rows.append([vid, primary_label, "", "",
                                     m["source"], m["src_id"], m["src_name"],
                                     m.get("src_lat") or "", m.get("src_lon") or "",
                                     m.get("src_class") or "",
                                     "", "ungeocoded_match"])
            else:
                counts["ungeocoded_no_match"] += 1
        else:
            if not hits:
                continue
            counts["geocoded_match"] += 1
            # Disagreement check against best-coord match
            for m in hits:
                slat, slon = m.get("src_lat"), m.get("src_lon")
                if slat is None or slon is None:
                    counts["geocoded_match_no_src_coord"] += 1
                    continue
                d = haversine_km(db_lat, db_lon, slat, slon)
                if d > args.disagreement_km:
                    counts["geocoded_disagree"] += 1
                    csv_rows.append([vid, primary_label, db_lat, db_lon,
                                     m["source"], m["src_id"], m["src_name"],
                                     slat, slon, m.get("src_class") or "",
                                     f"{d:.1f}", "geocoded_disagree"])

    counts["ungeocoded_unique_vocab"] = len(matched_vocab_ids)

    csv_path = out / "candidates.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["vocab_id", "label", "db_lat", "db_lon",
                    "source", "src_id", "src_name", "src_lat", "src_lon",
                    "src_class", "distance_km", "bucket"])
        w.writerows(csv_rows)

    print()
    print("=" * 64)
    print("PROBE SUMMARY")
    print("=" * 64)
    for k, v in counts.items():
        print(f"  {k:32s} {v:>8,}")
    print()
    print(f"Distinct ungeocoded vocab matches: {counts['ungeocoded_unique_vocab']:,}")
    decision = ("≥30 — RECOMMEND phase 1f for v0.26"
                if counts["ungeocoded_unique_vocab"] >= 30
                else "<30 — DOWNGRADE to label enrichment only")
    print(f"Decision rule: {decision}")
    print()
    print(f"CSV written to: {csv_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
